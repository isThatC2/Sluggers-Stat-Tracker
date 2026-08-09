from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from MemoryHandling.sluggers_data import Field, Player, Team, Game, NO_PLAYER, NO_TEAM
import datetime
from pathlib import Path
import sys


def _resource_path(filename: str) -> Path:
    """Resolve a release asset from beside the executable or its bundle."""
    relative_path = Path(filename)
    if relative_path.is_absolute():
        return relative_path

    if getattr(sys, "frozen", False):
        external_path = Path(sys.executable).resolve().parent / relative_path
        if external_path.exists():
            return external_path

        bundle_dir = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
        return bundle_dir / relative_path

    return Path(__file__).resolve().parent / relative_path


class Outputter:
    def __init__(self, template_filename: str, game: Game):
        self.tmp_filename = _resource_path(template_filename)
        self.game = game
        self.output_path = Path.cwd() / "output"
        self.output_path.mkdir(exist_ok=True)
        try:
            self.template = load_workbook(self.tmp_filename)
        except FileNotFoundError:
            raise FileNotFoundError(f"Template file '{self.tmp_filename}' not found.")
    
    
    def make_output_filename(self, game: Game) -> str:
        output_name = f"{game.away_team.short_name} vs {game.home_team.short_name} - {datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
        if Path(self.output_path / output_name).exists():
            output_name = f"{game.away_team.short_name} vs {game.home_team.short_name} - {datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')}_{datetime.datetime.now().microsecond}.xlsx"
        return output_name
    def save(self, output_filename: str):
        self.template.save(self.output_path / output_filename)
    
    
    def output_game_info(self, game_info_sheet, game: Game):
        team1 = game.away_team
        team2 = game.home_team
        gi = game_info_sheet
        if game.stat_tracker_started_during_match:
            gi["L9"] = "Stat Tracker Started During Match"
        else:
            gi["L9"] = None
        
        if game.match_quit_early:
            gi["L10"] = "Match Quit Early"
        else:
            gi["L10"] = None 
        
        gi["L11"] = team1.name
        gi["P11"] = team2.name
        
        gi["I12"] = team1.score.value
        gi["P12"] = team2.score.value
        
        gi["I13"] = team1.player_type.display
        gi["P13"] = team2.player_type.display
        
        
        gi["I14"] = f"{game.stadium.display} - {game.time_of_day.display}"
        gi["M15"] = f"Innings - {game.regulation_innings.value}"
        gi["M16"] = f"Mercy - {"On" if game.mercy_flag.value == 1 else "Off"}"
        gi["M17"] = f"Stars - {"On" if game.stars_flag.value == 1 else "Off"}"
        gi["M18"] = f"Items - {"On" if game.item_flag.value == 1 else "Off"}"
        
        starting_row = 20
        for i, player in enumerate(team1.players):
            gi[f"J{starting_row + i}"] = player.name
            if team1.starting_lineup_set:
                gi[f"K{starting_row + i}"] = team1.starting_lineup.get(player, "None")
            else:
                gi[f"K{starting_row + i}"] = "None"
        
        for i, player in enumerate(team2.players):
            gi[f"R{starting_row + i}"] = player.name
            if team2.starting_lineup_set:
                gi[f"Q{starting_row + i}"] = team2.starting_lineup.get(player, "N/A")
            else:
                gi[f"Q{starting_row + i}"] = "N/A"
        
        scoreboard_row = 30
        scoreboard_start_col = "K"
        scoreboard_col_idx = column_index_from_string(scoreboard_start_col)
        current_col = scoreboard_col_idx
        gi["J31"] = game.away_team.name
        gi["J32"] = game.home_team.name
        
        if game.current_inning.value > 1:
            for i in range(game.current_inning.value):
                gi.cell(row=scoreboard_row, column=current_col).value = i + 1
                gi.cell(row=scoreboard_row + 1, column=current_col).value = game.away_team.score_by_inning[i]
                gi.cell(row=scoreboard_row + 2, column=current_col).value = game.home_team.score_by_inning[i]

                if i + 1 > game.regulation_innings.value:
                    gi.cell(row=scoreboard_row, column=current_col).font = Font(color="FF0000")
                    gi.cell(row=scoreboard_row + 1, column=current_col).font = Font(color="FF0000")
                    gi.cell(row=scoreboard_row + 2, column=current_col).font = Font(color="FF0000")
                current_col += 1
        else:
            gi.cell(row=scoreboard_row, column=current_col).value = i + 1
            gi.cell(row=scoreboard_row + 1, column=current_col).value = game.away_team.score.value
            gi.cell(row=scoreboard_row + 2, column=current_col).value = game.home_team.score.value
            
        
        gi.cell(row=scoreboard_row, column=current_col).value = "Final"
        gi.cell(row=scoreboard_row, column=current_col).font = Font(bold=True)
        gi.cell(row=scoreboard_row, column=current_col).alignment = Alignment(horizontal='right')
        
        gi.cell(row=scoreboard_row + 1, column=current_col).value = game.away_team.score.value
        gi.cell(row=scoreboard_row + 1, column=current_col).font = Font(bold=True)
        gi.cell(row=scoreboard_row + 2, column=current_col).value = game.home_team.score.value
        gi.cell(row=scoreboard_row + 2, column=current_col).font = Font(bold=True)
        
            
            
            
        
        
        
    
    def output_player_stats_row(self, stats_sheet, player: Player, team: Team, row: int):
            header_columns = self._sheet_header_columns(stats_sheet)
            for header, get_stat in self._stats_stat_specs():
                if header not in header_columns:
                    continue
                stats_sheet.cell(row=row, column=header_columns[header], value=get_stat(player))
            
        
    def output_pitching_row(self, pitching_sheet, player: Player, team: Team, row: int):
        header_columns = self._pitching_header_columns(pitching_sheet)
        for header, get_stat, _ in self._pitching_stat_specs():
            pitching_sheet.cell(row=row, column=header_columns[header], value=get_stat(player, team))
    
    def output_stats(self, stats_sheet, game: Game, team1: Team, team2: Team):
        header_columns = self._sheet_header_columns(stats_sheet)
        self._ensure_stats_helper_columns(stats_sheet, header_columns)
        header_columns = self._sheet_header_columns(stats_sheet)

        stats_sheet.cell(row=2, column=header_columns["Team"], value=team1.short_name)
        stats_sheet.cell(row=12, column=header_columns["Team"], value=team2.short_name)

        for i, player in enumerate(team1.players):
            self.output_player_stats_row(stats_sheet, player, team1, i + 2)
        
        for i, player in enumerate(team2.players):
            self.output_player_stats_row(stats_sheet, player, team2, i + 12)

        self.output_stats_team_totals(stats_sheet, team1, 11, 2, 10, header_columns, fill_color="#F4CCCC")
        self.output_stats_team_totals(stats_sheet, team2, 21, 12, 20, header_columns, fill_color="#C9DAF8")
            
            
            
            
        
        
    
    def output_pitching(self, pitching_sheet, game: Game, team1: Team, team2: Team):
        header_columns = self._sheet_header_columns(pitching_sheet)
        self._ensure_pitching_helper_columns(pitching_sheet, header_columns)
        header_columns = self._sheet_header_columns(pitching_sheet)
        row_num = 2
        team1_start_row = row_num
        for player in team1.pitcher_order:
            self.output_pitching_row(pitching_sheet, player, team1, row_num)
            row_num += 1

        team2_start_row = row_num
        for player in team2.pitcher_order:
            self.output_pitching_row(pitching_sheet, player, team2, row_num)
            row_num += 1

        self.output_pitching_team_totals(
            pitching_sheet,
            team1,
            row_num,
            team1_start_row,
            header_columns,
            fill_color="#F4CCCC",
        )
        row_num += 1
        self.output_pitching_team_totals(
            pitching_sheet,
            team2,
            row_num,
            team2_start_row,
            header_columns,
            fill_color="#C9DAF8",
        )

    def output_pitching_team_totals(
        self,
        pitching_sheet,
        team: Team,
        row: int,
        team_start_row: int,
        header_columns: dict[str, int],
        fill_color: str | None = None,
    ):
        num_rows_to_read = len(team.pitcher_order)
        start_row = team_start_row
        end_row = team_start_row + num_rows_to_read - 1

        pitching_sheet.cell(row=row, column=1, value=f"{team.short_name} Totals")
        pitching_sheet.cell(row=row, column=2, value=team.short_name)

        for header, _, aggregate_type in self._pitching_stat_specs():
            if aggregate_type is None:
                continue
            col_idx = header_columns[header]
            formula = self._pitching_team_total_formula(header, aggregate_type, row, start_row, end_row, header_columns)
            pitching_sheet.cell(row=row, column=col_idx, value=formula)

        if fill_color:
            normalized = fill_color.lstrip("#")
            if len(normalized) == 6:
                normalized = f"FF{normalized}"
            fill = PatternFill(fill_type="solid", start_color=normalized, end_color=normalized)
            for col_idx in range(1, max(header_columns.values()) + 1):
                pitching_sheet.cell(row=row, column=col_idx).fill = fill

    def output_stats_team_totals(
        self,
        stats_sheet,
        team: Team,
        row: int,
        team_start_row: int,
        team_end_row: int,
        header_columns: dict[str, int],
        fill_color: str | None = None,
    ):
        
        stats_sheet.cell(row=row, column=header_columns["Player"], value=team.short_name)
        if "Position" in header_columns:
            stats_sheet.cell(row=row, column=header_columns["Position"], value="N/A")

        for header, aggregate_type in self._stats_total_specs():
            if header not in header_columns:
                continue
            col_idx = header_columns[header]
            formula = self._stats_team_total_formula(
                aggregate_type,
                header,
                row,
                team_start_row,
                team_end_row,
                header_columns,
            )
            stats_sheet.cell(row=row, column=col_idx, value=formula)

        if fill_color:
            normalized = fill_color.lstrip("#")
            if len(normalized) == 6:
                normalized = f"FF{normalized}"
            fill = PatternFill(fill_type="solid", start_color=normalized, end_color=normalized)
            for col_idx in range(1, max(header_columns.values()) + 1):
                stats_sheet.cell(row=row, column=col_idx).fill = fill

    def _pitching_stat_specs(self):
        def check_inf(val):
            return "INF" if val == float('inf') else val

        # (header, value_getter, team_totals_aggregation)
        return [
            ("Player", lambda p, t: p.name, None),
            ("Team", lambda p, t: t.short_name, None),
            ("Batters Faced", lambda p, t: p.stats.pitching.batters_faced, "sum"),
            ("Innings Pitched", lambda p, t: p.stats.pitching.innings_pitched, "sum"),
            ("Pitches", lambda p, t: p.stats.pitching.pitch_count, "sum"),
            ("Strikes", lambda p, t: p.stats.pitching.strikes, "sum"),
            ("Balls", lambda p, t: p.stats.pitching.balls, "sum"),
            ("Strikeouts", lambda p, t: p.stats.pitching.strikeouts, "sum"),
            ("Walks", lambda p, t: p.stats.pitching.walks, "sum"),
            ("Bean Balls", lambda p, t: p.stats.pitching.bean_balls, "sum"),
            ("Hits Allowed", lambda p, t: p.stats.pitching.hits_allowed, "sum"),
            ("Runs Allowed", lambda p, t: p.stats.pitching.runs_allowed, "sum"),
            ("Singles Allowed", lambda p, t: p.stats.pitching.singles_allowed, "sum"),
            ("Doubles Allowed", lambda p, t: p.stats.pitching.doubles_allowed, "sum"),
            ("Triples Allowed", lambda p, t: p.stats.pitching.triples_allowed, "sum"),
            ("HR Allowed", lambda p, t: p.stats.pitching.home_runs_allowed, "sum"),
            ("Earned Runs", lambda p, t: p.stats.pitching.earned_runs, "sum"),
            ("Inherited Runs", lambda p, t: p.stats.pitching.inherited_runs, "sum"),
            ("Star Pitches", lambda p, t: p.stats.pitching.star_pitches, "sum"),
            ("Stars Used", lambda p, t: p.stats.pitching.stars_used, "sum"),
            ("Pickoffs", lambda p, t: p.stats.pitching.pickoffs, "sum"),
            ("Pickoff Attempts", lambda p, t: p.stats.pitching.pickoff_attempts, "sum"),
            ("ERA-7", lambda p, t: check_inf(p.stats.pitching.era_per_7), "era_7"),
            ("ERA-9", lambda p, t: check_inf(p.stats.pitching.era_per_9), "era_9"),
            ("WHIP", lambda p, t: check_inf(p.stats.pitching.whip), "whip"),
            ("BA Against", lambda p, t: check_inf(p.stats.pitching.batting_average_against), "batting_average_against"),
            ("OB% Against", lambda p, t: check_inf(p.stats.pitching.on_base_percentage_against), "on_base_percentage_against"),
            ("SLG Against", lambda p, t: check_inf(p.stats.pitching.slugging_percentage_against), "slugging_percentage_against"),
            ("OPS Against", lambda p, t: check_inf(p.stats.pitching.on_base_slugging_against), "on_base_slugging_against"),
            ("At-Bats Against Helper", lambda p, t: p.stats.pitching.at_bats_against, None),
            ("Total Bases Allowed Helper", lambda p, t: p.stats.pitching.total_bases_allowed, None),
        ]

    def _stats_stat_specs(self):
        return [
            ("Player", lambda p: p.name),
            ("Position", lambda p: ", ".join(p.stats.positions_played)),
            ("At-Bats", lambda p: p.stats.batting.at_bats),
            ("Plate Appearances", lambda p: p.stats.batting.plate_appearances),
            ("Runs", lambda p: p.stats.running.runs),
            ("Hits", lambda p: p.stats.batting.hits),
            ("RBI", lambda p: p.stats.batting.rbi),
            ("Strikeouts", lambda p: p.stats.batting.strikeouts),
            ("Walks", lambda p: p.stats.batting.walks),
            ("Hit By Pitch", lambda p: p.stats.batting.hit_by_pitch),
            ("Singles", lambda p: p.stats.batting.singles),
            ("Doubles", lambda p: p.stats.batting.doubles),
            ("Triples", lambda p: p.stats.batting.triples),
            ("Home Runs", lambda p: p.stats.batting.home_runs),
            ("1HR", lambda p: p.stats.batting.one_run_homeruns),
            ("2HR", lambda p: p.stats.batting.two_run_homeruns),
            ("3HR", lambda p: p.stats.batting.three_run_homeruns),
            ("Grand Slams", lambda p: p.stats.batting.grand_slams),
            ("ITP Home Runs", lambda p: p.stats.batting.inside_the_park_home_runs),
            ("Total Bases", lambda p: p.stats.batting.total_bases),
            ("Sac Flys", lambda p: p.stats.batting.sac_flys),
            ("Star Hits", lambda p: p.stats.batting.star_hits),
            ("Stars Used", lambda p: p.stats.batting.stars_used),
            ("Batting Average", lambda p: p.stats.batting.batting_average),
            ("On Base %", lambda p: p.stats.batting.on_base_percentage),
            ("Slug %", lambda p: p.stats.batting.slugging_percentage),
            ("On Base + Slug", lambda p: p.stats.batting.on_base_slugging),
            ("Star Slug %", lambda p: p.stats.batting.star_slugging_percentage),
            ("Stolen Bases", lambda p: p.stats.running.stolen_bases),
            ("Caught Stealing", lambda p: p.stats.running.caught_stealing),
            ("Steal Attempts", lambda p: p.stats.running.steal_attempts),
            ("Putouts", lambda p: p.stats.fielding.putouts),
            ("Assists", lambda p: p.stats.fielding.assists),
            ("Buddy Jump Putouts", lambda p: p.stats.fielding.buddy_jump_outs),
            ("Buddy Jump Attempts", lambda p: p.stats.fielding.buddy_jump_attempts),
            ("Double Plays", lambda p: p.stats.fielding.double_plays),
            ("Triple Plays", lambda p: p.stats.fielding.triple_plays),
            ("Bobbles", lambda p: p.stats.fielding.bobbles),
            ("Star Bases Helper", lambda p: p.stats.batting.total_star_bases),
        ]

    def _stats_total_specs(self):
        return [
            ("At-Bats", "sum"),
            ("Plate Appearances", "sum"),
            ("Runs", "sum"),
            ("Hits", "sum"),
            ("RBI", "sum"),
            ("Strikeouts", "sum"),
            ("Walks", "sum"),
            ("Hit By Pitch", "sum"),
            ("Singles", "sum"),
            ("Doubles", "sum"),
            ("Triples", "sum"),
            ("Home Runs", "sum"),
            ("1HR", "sum"),
            ("2HR", "sum"),
            ("3HR", "sum"),
            ("Grand Slams", "sum"),
            ("ITP Home Runs", "sum"),
            ("Total Bases", "sum"),
            ("Sac Flys", "sum"),
            ("Star Hits", "sum"),
            ("Stars Used", "sum"),
            ("Batting Average", "batting_average"),
            ("On Base %", "on_base_percentage"),
            ("Slug %", "slugging_percentage"),
            ("On Base + Slug", "on_base_slugging"),
            ("Star Slug %", "star_slugging_percentage"),
            ("Stolen Bases", "sum"),
            ("Caught Stealing", "sum"),
            ("Steal Attempts", "sum"),
            ("Putouts", "sum"),
            ("Assists", "sum"),
            ("Buddy Jump Putouts", "sum"),
            ("Buddy Jump Attempts", "sum"),
            ("Double Plays", "sum"),
            ("Triple Plays", "sum"),
            ("Bobbles", "sum"),
        ]

    def _pitching_header_columns(self, pitching_sheet) -> dict[str, int]:
        return self._sheet_header_columns(pitching_sheet)

    def _sheet_header_columns(self, sheet) -> dict[str, int]:
        header_columns = {}
        for col_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col_idx).value
            if header:
                header_columns[str(header)] = col_idx
        return header_columns

    def _ensure_stats_helper_columns(self, stats_sheet, header_columns: dict[str, int]):
        if "Star Bases Helper" in header_columns:
            return

        helper_col = stats_sheet.max_column + 1
        stats_sheet.cell(row=1, column=helper_col, value="Star Bases Helper")
        stats_sheet.column_dimensions[get_column_letter(helper_col)].hidden = True

    def _ensure_pitching_helper_columns(self, pitching_sheet, header_columns: dict[str, int]):
        helper_headers = ["At-Bats Against Helper", "Total Bases Allowed Helper"]
        missing_helpers = [header for header in helper_headers if header not in header_columns]
        if not missing_helpers:
            return

        next_col = pitching_sheet.max_column + 1
        for header in missing_helpers:
            pitching_sheet.cell(row=1, column=next_col, value=header)
            pitching_sheet.column_dimensions[get_column_letter(next_col)].hidden = True
            next_col += 1

    def _sum_formula(self, column_letter: str, start_row: int, end_row: int) -> str:
        return f"=SUM({column_letter}{start_row}:{column_letter}{end_row})"

    def _ratio_formula(self, numerator: str, denominator: str, row: int, zero_value: str = "0") -> str:
        return f'=IF({denominator}{row}=0,{zero_value},{numerator}{row}/{denominator}{row})'

    def _stats_team_total_formula(
        self,
        aggregate_type: str,
        header: str,
        row: int,
        team_start_row: int,
        team_end_row: int,
        header_columns: dict[str, int],
    ) -> str:
        handlers = {
            "sum": lambda: self._sum_formula(get_column_letter(header_columns[header]), team_start_row, team_end_row),
            "batting_average": lambda: self._stats_batting_average_formula(row, header_columns),
            "on_base_percentage": lambda: self._stats_on_base_percentage_formula(row, header_columns),
            "slugging_percentage": lambda: self._stats_slugging_formula(row, header_columns),
            "on_base_slugging": lambda: self._stats_on_base_slugging_formula(row, header_columns),
            "star_slugging_percentage": lambda: self._stats_star_slugging_formula(row, team_start_row, team_end_row, header_columns),
        }

        if aggregate_type not in handlers:
            raise ValueError(f"Unsupported stats aggregate type: {aggregate_type}")

        return handlers[aggregate_type]()

    def _pitching_team_total_formula(
        self,
        header: str,
        aggregate_type: str,
        row: int,
        team_start_row: int,
        team_end_row: int,
        header_columns: dict[str, int],
    ) -> str:
        handlers = {
            "sum": lambda: self._sum_formula(get_column_letter(header_columns[header]), team_start_row, team_end_row),
            "era_7": lambda: self._pitching_era_formula(row, header_columns, 7),
            "era_9": lambda: self._pitching_era_formula(row, header_columns, 9),
            "whip": lambda: self._pitching_whip_formula(row, header_columns),
            "batting_average_against": lambda: self._pitching_batting_average_against_formula(row, team_start_row, team_end_row, header_columns),
            "on_base_percentage_against": lambda: self._pitching_on_base_percentage_against_formula(row, team_start_row, team_end_row, header_columns),
            "slugging_percentage_against": lambda: self._pitching_slugging_percentage_against_formula(row, team_start_row, team_end_row, header_columns),
            "on_base_slugging_against": lambda: self._pitching_on_base_slugging_against_formula(row, header_columns),
        }

        if aggregate_type not in handlers:
            raise ValueError(f"Unsupported pitching aggregate type: {aggregate_type}")

        return handlers[aggregate_type]()

    def _stats_batting_average_formula(self, row: int, header_columns: dict[str, int]) -> str:
        hits_col = get_column_letter(header_columns["Hits"])
        at_bats_col = get_column_letter(header_columns["At-Bats"])
        return f'=IF({at_bats_col}{row}=0,0,{hits_col}{row}/{at_bats_col}{row})'

    def _stats_on_base_percentage_formula(self, row: int, header_columns: dict[str, int]) -> str:
        hits_col = get_column_letter(header_columns["Hits"])
        walks_col = get_column_letter(header_columns["Walks"])
        hbp_col = get_column_letter(header_columns["Hit By Pitch"])
        pa_col = get_column_letter(header_columns["Plate Appearances"])
        return f'=IF({pa_col}{row}=0,0,({hits_col}{row}+{walks_col}{row}+{hbp_col}{row})/{pa_col}{row})'

    def _stats_slugging_formula(self, row: int, header_columns: dict[str, int]) -> str:
        total_bases_col = get_column_letter(header_columns["Total Bases"])
        at_bats_col = get_column_letter(header_columns["At-Bats"])
        return f'=IF({at_bats_col}{row}=0,0,{total_bases_col}{row}/{at_bats_col}{row})'

    def _stats_on_base_slugging_formula(self, row: int, header_columns: dict[str, int]) -> str:
        obp_col = get_column_letter(header_columns["On Base %"])
        slg_col = get_column_letter(header_columns["Slug %"])
        return f'={obp_col}{row}+{slg_col}{row}'

    def _stats_star_slugging_formula(self, row: int, team_start_row: int, team_end_row: int, header_columns: dict[str, int]) -> str:
        star_bases_col = get_column_letter(header_columns["Star Bases Helper"])
        stars_used_col = get_column_letter(header_columns["Stars Used"])
        return f'=IF({stars_used_col}{row}=0,0,SUM({star_bases_col}{team_start_row}:{star_bases_col}{team_end_row})/{stars_used_col}{row})'

    def _pitching_era_formula(self, row: int, header_columns: dict[str, int], innings_factor: int) -> str:
        earned_runs_col = get_column_letter(header_columns["Earned Runs"])
        innings_col = get_column_letter(header_columns["Innings Pitched"])
        return f'=IF({innings_col}{row}=0,"INF",{earned_runs_col}{row}*{innings_factor}/{innings_col}{row})'

    def _pitching_whip_formula(self, row: int, header_columns: dict[str, int]) -> str:
        walks_col = get_column_letter(header_columns["Walks"])
        hits_col = get_column_letter(header_columns["Hits Allowed"])
        innings_col = get_column_letter(header_columns["Innings Pitched"])
        return f'=IF({innings_col}{row}=0,"INF",({walks_col}{row}+{hits_col}{row})/{innings_col}{row})'

    def _pitching_batting_average_against_formula(self, row: int, team_start_row: int, team_end_row: int, header_columns: dict[str, int]) -> str:
        hits_col = get_column_letter(header_columns["Hits Allowed"])
        at_bats_helper_col = get_column_letter(header_columns["At-Bats Against Helper"])
        return f'=IF(SUM({at_bats_helper_col}{team_start_row}:{at_bats_helper_col}{team_end_row})=0,0,SUM({hits_col}{team_start_row}:{hits_col}{team_end_row})/SUM({at_bats_helper_col}{team_start_row}:{at_bats_helper_col}{team_end_row}))'

    def _pitching_on_base_percentage_against_formula(self, row: int, team_start_row: int, team_end_row: int, header_columns: dict[str, int]) -> str:
        hits_col = get_column_letter(header_columns["Hits Allowed"])
        walks_col = get_column_letter(header_columns["Walks"])
        bean_balls_col = get_column_letter(header_columns["Bean Balls"])
        batters_faced_col = get_column_letter(header_columns["Batters Faced"])
        return f'=IF(SUM({batters_faced_col}{team_start_row}:{batters_faced_col}{team_end_row})=0,0,(SUM({hits_col}{team_start_row}:{hits_col}{team_end_row})+SUM({walks_col}{team_start_row}:{walks_col}{team_end_row})+SUM({bean_balls_col}{team_start_row}:{bean_balls_col}{team_end_row}))/SUM({batters_faced_col}{team_start_row}:{batters_faced_col}{team_end_row}))'

    def _pitching_slugging_percentage_against_formula(self, row: int, team_start_row: int, team_end_row: int, header_columns: dict[str, int]) -> str:
        total_bases_helper_col = get_column_letter(header_columns["Total Bases Allowed Helper"])
        at_bats_helper_col = get_column_letter(header_columns["At-Bats Against Helper"])
        return f'=IF(SUM({at_bats_helper_col}{team_start_row}:{at_bats_helper_col}{team_end_row})=0,0,SUM({total_bases_helper_col}{team_start_row}:{total_bases_helper_col}{team_end_row})/SUM({at_bats_helper_col}{team_start_row}:{at_bats_helper_col}{team_end_row}))'

    def _pitching_on_base_slugging_against_formula(self, row: int, header_columns: dict[str, int]) -> str:
        obp_col = get_column_letter(header_columns["OB% Against"])
        slg_col = get_column_letter(header_columns["SLG Against"])
        return f'={obp_col}{row}+{slg_col}{row}'


    def output_game(self, game: Game) -> str:
        if "Stats" not in self.template.sheetnames:
            raise ValueError("Template does not contain a 'Stats' sheet.")
        
        if "Pitching" not in self.template.sheetnames:
            raise ValueError("Template does not contain a 'Pitching' sheet.")
        
        game_info_sheet = self.template["Game Info"]
        stats_sheet = self.template["Stats"]
        pitching_sheet = self.template["Pitching"]
        
        print("Beginning output of game data to Excel...")
        self.output_game_info(game_info_sheet, game)
        self.output_stats(stats_sheet, game, game.away_team, game.home_team)
        self.output_pitching(pitching_sheet, game, game.away_team, game.home_team)
        print("Saving output file...")
        output_filename = self.make_output_filename(game)
        self.save(output_filename)
        print(f"Output saved to '{output_filename}'")
        return output_filename
